"""Generate sample images and Excel seed files for LabTrack."""
import os
import random
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
IMAGES_DIR = os.path.join(DATA_DIR, 'images')


def make_image(filename, width=400, height=300, bg_color=None, label=""):
    """Generate a simple colored rectangle image with a label."""
    if bg_color is None:
        bg_color = (
            random.randint(60, 200),
            random.randint(60, 200),
            random.randint(60, 200),
        )
    img = Image.new('RGB', (width, height), bg_color)
    draw = ImageDraw.Draw(img)
    # Try to use a default font
    try:
        font = ImageFont.truetype("arial.ttf", 20)
    except OSError:
        font = ImageFont.load_default()
    # Draw a border
    draw.rectangle([5, 5, width - 5, height - 5], outline=(255, 255, 255), width=3)
    # Draw label
    if label:
        bbox = draw.textbbox((0, 0), label, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        x = (width - text_w) // 2
        y = (height - text_h) // 2
        draw.text((x, y), label, fill=(255, 255, 255), font=font)
    img.save(filename)
    return filename


# ── Generate sample images ───────────────────────────────────────────────

def generate_images():
    print("Generating sample images...")

    equipment_items = [
        ("Oscilloscope", (100, 50, 50)),
        ("Multimeter", (50, 100, 50)),
        ("Arduino", (50, 50, 100)),
        ("Raspberry Pi", (150, 80, 30)),
        ("3D Printer", (80, 30, 150)),
        ("Laser Cutter", (30, 120, 80)),
        ("Signal Gen", (120, 60, 60)),
        ("Spectrum", (60, 120, 120)),
        ("Soldering", (120, 120, 60)),
        ("Logic Analyzer", (80, 80, 80)),
        ("Microscope", (60, 60, 120)),
        ("Power Supply", (100, 80, 40)),
    ]
    for i, (name, color) in enumerate(equipment_items):
        make_image(
            os.path.join(IMAGES_DIR, 'equipment', f'eq_{i+1:02d}.png'),
            label=name,
            bg_color=color,
        )

    profiles = [
        ("Dr. Tanzeel", (70, 130, 180)),
        ("Hasibur", (180, 70, 70)),
        ("Mohtasim", (70, 180, 70)),
        ("Hamid", (180, 130, 70)),
        ("Waseem", (130, 70, 180)),
        ("Faraz", (70, 180, 180)),
        ("Maruf", (180, 70, 130)),
        ("Rohan", (130, 180, 70)),
    ]
    for i, (name, color) in enumerate(profiles):
        make_image(
            os.path.join(IMAGES_DIR, 'profiles', f'profile_{i+1:02d}.png'),
            width=200, height=200,
            label=name,
            bg_color=color,
        )

    incidents = [
        ("Broken Screen", (200, 50, 50)),
        ("Cable Damage", (200, 100, 50)),
    ]
    for i, (name, color) in enumerate(incidents):
        make_image(
            os.path.join(IMAGES_DIR, 'incidents', f'incident_{i+1:02d}.png'),
            label=name,
            bg_color=color,
        )

    print(f"  Images created in {IMAGES_DIR}")


# ── Excel helpers ────────────────────────────────────────────────────────

def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border


def create_users_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    headers = ["username", "email", "first_name", "last_name", "role", "password", "image_path"]
    ws.append(headers)
    rows = [
        ["tur0001", "tur0001@auburn.edu", "Dr. Tanzeel", "Rehman", "ADMIN", "admin123", "profiles/profile_01.png"],
        ["mzr0134", "mzr0134@auburn.edu", "Md Hasibur", "Rahman", "ADMIN", "admin123", "profiles/profile_02.png"],
        ["mzr0167", "mzr0167@auburn.edu", "Mohtasim Hadi", "Rafi", "ADMIN", "admin123", "profiles/profile_03.png"],
        ["hhs0015", "hhs0015@auburn.edu", "Hamid Habib", "Syed", "MEMBER", "member123", "profiles/profile_04.png"],
        ["mzw0147", "mzw0147@auburn.edu", "Mohammad", "Waseem", "MEMBER", "member123", "profiles/profile_05.png"],
        ["fza0070", "fza0070@auburn.edu", "Faraz", "Ahmad", "MEMBER", "member123", "profiles/profile_06.png"],
        ["mzm0411", "mzm0411@auburn.edu", "Md Mesbahul", "Maruf", "MEMBER", "member123", "profiles/profile_07.png"],
        ["mza0291", "mza0291@auburn.edu", "Rohan", "Afzal", "MEMBER", "member123", "profiles/profile_08.png"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'users.xlsx'))
    print("  Created users.xlsx")


def create_categories_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    headers = ["name", "description", "color"]
    ws.append(headers)
    rows = [
        ["Jetson", "NVIDIA Jetson edge AI devices", ""],
        ["Desktop Computer", "Desktop workstations and PCs", ""],
        ["Laptop Computer", "Portable laptops and notebooks", ""],
        ["Edge Device", "IoT and edge computing devices", ""],
        ["Router", "Network routers and switches", ""],
        ["Drone", "UAVs and flying robots", ""],
        ["Camera", "Cameras and imaging devices", ""],
        ["Storage Device", "HDDs, SSDs, and NAS units", ""],
        ["Measurement", "Test and measurement equipment", ""],
        ["Power", "Power supplies and batteries", ""],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'categories.xlsx'))
    print("  Created categories.xlsx")


def create_locations_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"
    headers = ["name", "description", "building", "room"]
    ws.append(headers)
    rows = [
        ["Main Lab", "Primary workspace with benches and power", "Engineering Building", "Room 101"],
        ["Storage Room", "Secure storage for spare parts and tools", "Engineering Building", "Room 105"],
        ["Electronics Bench", "Soldering and prototyping area", "Engineering Building", "Room 102"],
        ["Clean Room", "Dust-free environment for sensitive work", "Science Center", "Room 201"],
        ["Workshop", "Heavy tools and fabrication space", "Engineering Building", "Room 110"],
        ["Server Room", "Rack-mounted servers and networking", "Engineering Building", "Room 115"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'locations.xlsx'))
    print("  Created locations.xlsx")


def create_equipment_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"
    headers = [
        "name", "description", "serial_number", "model_number", "manufacturer",
        "category_name", "location_name", "owner_username", "condition", "status",
        "image_path", "purchase_date", "purchase_price", "notes"
    ]
    ws.append(headers)
    rows = [
        ["Oscilloscope Tektronix MSO44", "Digital oscilloscope with 4 channels, 200MHz bandwidth", "SN-OSC-001", "MSO44", "Tektronix", "Measurement", "Electronics Bench", "tur0001", "EXCELLENT", "AVAILABLE", "equipment/eq_01.png", "2023-01-15", "4500.00", "Calibrated quarterly"],
        ["Multimeter Fluke 87V", "Professional industrial multimeter with true RMS", "SN-MMT-001", "87V", "Fluke", "Measurement", "Electronics Bench", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_02.png", "2022-08-10", "350.00", ""],
        ["Arduino Mega 2560", "Microcontroller development board with 54 I/O pins", "SN-ARD-001", "A000067", "Arduino", "Edge Device", "Main Lab", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_03.png", "2023-03-20", "45.00", ""],
        ["Raspberry Pi 4 (8GB)", "Single-board computer with 8GB RAM and quad-core CPU", "SN-RPI-001", "SC0195", "Raspberry Pi Foundation", "Edge Device", "Main Lab", "tur0001", "EXCELLENT", "AVAILABLE", "equipment/eq_04.png", "2023-06-01", "75.00", "In protective case"],
        ["3D Printer - Prusa MK4", "FDM 3D printer with auto-bed leveling", "SN-3DP-001", "MK4", "Prusa Research", "Drone", "Workshop", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_05.png", "2023-09-12", "1100.00", "PLA and PETG profiles configured"],
        ["Laser Cutter 60W", "CO2 laser cutter/engraver with rotary attachment", "SN-LC-001", "xTool P2", "xTool", "Drone", "Workshop", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_06.png", "2023-11-05", "3200.00", "Ventilation required"],
        ["Signal Generator", "Function/arbitrary waveform generator up to 25MHz", "SN-SIG-001", "DG1022Z", "Rigol", "Measurement", "Electronics Bench", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_07.png", "2022-12-01", "280.00", ""],
        ["Spectrum Analyzer", "RF spectrum analyzer 9kHz-3GHz", "SN-SPA-001", "DSA815", "Rigol", "Measurement", "Electronics Bench", "tur0001", "EXCELLENT", "AVAILABLE", "equipment/eq_08.png", "2023-02-20", "1200.00", ""],
        ["Soldering Station", "Temperature-controlled soldering station with hot air", "SN-SOL-001", "FX-888D", "Hakko", "Measurement", "Electronics Bench", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_09.png", "2022-05-15", "120.00", "Extra tips in drawer"],
        ["Logic Analyzer", "16-channel logic analyzer with protocol decoding", "SN-LA-001", "Logic Pro 16", "Saleae", "Measurement", "Electronics Bench", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_10.png", "2023-04-10", "400.00", ""],
        ["Digital Microscope", "USB digital microscope with 1000x magnification", "SN-MIC-001", "MD800", "AmScope", "Camera", "Main Lab", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_11.png", "2023-07-22", "180.00", ""],
        ["Power Supply (Bench)", "Programmable DC power supply 0-30V/5A dual channel", "SN-PSU-001", "DP832", "Rigol", "Power", "Electronics Bench", "tur0001", "GOOD", "AVAILABLE", "equipment/eq_12.png", "2022-10-30", "350.00", ""],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'equipment.xlsx'))
    print("  Created equipment.xlsx")


def create_kits_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kits"
    headers = ["name", "description", "created_by_username"]
    ws.append(headers)
    rows = [
        ["Electronics Starter Kit", "Complete kit for electronics prototyping with breadboard and components", "tur0001"],
        ["IoT Sensor Kit", "Collection of temperature, humidity, and motion sensors for IoT projects", "tur0001"],
        ["Robotics Kit", "Motors, motor drivers, and chassis parts for robot building", "tur0001"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'kits.xlsx'))
    print("  Created kits.xlsx")


def create_kit_items_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kit Items"
    headers = ["kit_name", "equipment_serial_number", "quantity"]
    ws.append(headers)
    rows = [
        ["Electronics Starter Kit", "SN-OSC-001", 1],
        ["Electronics Starter Kit", "SN-MMT-001", 1],
        ["Electronics Starter Kit", "SN-SOL-001", 1],
        ["IoT Sensor Kit", "SN-ARD-001", 1],
        ["IoT Sensor Kit", "SN-RPI-001", 1],
        ["Robotics Kit", "SN-ARD-001", 2],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'kit_items.xlsx'))
    print("  Created kit_items.xlsx")


def create_projects_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = ["name", "description", "lead_username", "status"]
    ws.append(headers)
    rows = [
        ["Smart Sensor Network", "IoT sensor network for environmental monitoring across campus", "hhs0015", "ACTIVE"],
        ["Autonomous Drone Fleet", "Multi-drone coordination for aerial surveillance and mapping", "mzw0147", "ACTIVE"],
        ["Edge AI Vision System", "Real-time object detection using Jetson and camera modules", "fza0070", "PLANNING"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'projects.xlsx'))
    print("  Created projects.xlsx")


def create_project_members_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Members"
    headers = ["project_name", "user_username", "role"]
    ws.append(headers)
    rows = [
        ["Smart Sensor Network", "hhs0015", "LEAD"],
        ["Smart Sensor Network", "mzw0147", "MEMBER"],
        ["Smart Sensor Network", "fza0070", "MEMBER"],
        ["Autonomous Drone Fleet", "mzw0147", "LEAD"],
        ["Autonomous Drone Fleet", "mzm0411", "MEMBER"],
        ["Edge AI Vision System", "fza0070", "LEAD"],
        ["Edge AI Vision System", "mza0291", "MEMBER"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'project_members.xlsx'))
    print("  Created project_members.xlsx")


def create_consumables_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumables"
    headers = ["name", "quantity", "low_stock_threshold", "unit", "category_name", "location_name", "unit_cost", "description"]
    ws.append(headers)
    rows = [
        ["Resistor Kit 1/4W", 600, 50, "PIECE", "Measurement", "Storage Room", 0.01, "Assorted 1/4W carbon film resistors"],
        ["Capacitor Kit (ceramic)", 400, 30, "PIECE", "Measurement", "Storage Room", 0.02, "Various ceramic capacitors"],
        ["Jumper Wires (M-M)", 100, 20, "PIECE", "Measurement", "Storage Room", 0.05, "Male-to-male breadboard jumper wires"],
        ["Solder Wire 60/40 500g", 8, 2, "ROLL", "Measurement", "Storage Room", 5.00, "Rosin-core solder wire"],
        ["Isopropyl Alcohol 99%", 5, 1, "BOTTLE", "Measurement", "Storage Room", 8.00, "For cleaning PCBs and components"],
        ["Safety Gloves (box)", 3, 1, "BOX", "Measurement", "Storage Room", 12.00, "Nitrile disposable gloves"],
        ["Thermal Paste", 4, 1, "PIECE", "Measurement", "Storage Room", 6.50, "High-performance CPU thermal paste"],
        ["PCB Prototype Boards", 25, 5, "PIECE", "Measurement", "Storage Room", 1.50, "Double-sided perfboard"],
        ["USB-C Cables", 15, 3, "PIECE", "Measurement", "Storage Room", 3.50, "1-meter USB-C to USB-A cables"],
        ["Breadboards (half)", 20, 4, "PIECE", "Measurement", "Storage Room", 2.00, "400-tie half-size breadboards"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'consumables.xlsx'))
    print("  Created consumables.xlsx")


def create_reservations_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations"
    headers = ["requester_username", "equipment_serial_number", "kit_name", "start_date", "end_date", "purpose", "status"]
    ws.append(headers)
    rows = [
        ["hhs0015", "SN-OSC-001", "", "2025-06-01", "2025-06-07", "Signal analysis for smart sensor project", "CONFIRMED"],
        ["mzw0147", "", "Robotics Kit", "2025-06-10", "2025-06-20", "Drone motor testing and assembly", "PENDING"],
        ["fza0070", "SN-RPI-001", "", "2025-07-01", "2025-07-15", "Jetson camera integration experiments", "CONFIRMED"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'reservations.xlsx'))
    print("  Created reservations.xlsx")


def create_borrows_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Borrows"
    headers = ["borrower_username", "equipment_serial_number", "kit_name", "due_date", "purpose", "status"]
    ws.append(headers)
    rows = [
        ["hhs0015", "SN-MMT-001", "", "2025-06-15", "Voltage measurement for sensor calibration", "APPROVED"],
        ["mzw0147", "", "IoT Sensor Kit", "2025-06-30", "Deploying sensors in the field", "ACTIVE"],
        ["mzm0411", "SN-3DP-001", "", "2025-07-10", "Printing drone propeller prototypes", "APPROVED"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    wb.save(os.path.join(DATA_DIR, 'borrows.xlsx'))
    print("  Created borrows.xlsx")


if __name__ == '__main__':
    generate_images()
    create_users_xlsx()
    create_categories_xlsx()
    create_locations_xlsx()
    create_equipment_xlsx()
    create_kits_xlsx()
    create_kit_items_xlsx()
    create_projects_xlsx()
    create_project_members_xlsx()
    create_consumables_xlsx()
    create_reservations_xlsx()
    create_borrows_xlsx()
    print("\nAll seed data files generated successfully!")

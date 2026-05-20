import os
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.core.files import File
from django.core.management import call_command
from django.core.management.base import BaseCommand
from django.db.utils import OperationalError, ProgrammingError
from django.utils import timezone
from openpyxl import load_workbook

User = __import__('django.contrib.auth', fromlist=['get_user_model']).get_user_model()


class Command(BaseCommand):
    help = 'Seed database with sample data from Excel files (auto-runs migrations if needed)'

    DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))), 'data')
    IMAGES_DIR = os.path.join(DATA_DIR, 'images')

    def _ensure_migrated(self):
        try:
            User.objects.exists()
        except (OperationalError, ProgrammingError):
            self.stdout.write(self.style.WARNING('Database tables missing. Running migrations first...'))
            call_command('migrate', interactive=False)
            self.stdout.write(self.style.SUCCESS('Migrations complete.'))

    def _load_sheet(self, filename):
        path = os.path.join(self.DATA_DIR, filename)
        if not os.path.exists(path):
            self.stdout.write(self.style.WARNING(f'  Missing {filename} — skipping.'))
            return []
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
        return rows

    def _load_image(self, image_path):
        """Return a Django File object from a relative image path."""
        if not image_path:
            return None
        full_path = os.path.join(self.IMAGES_DIR, image_path.replace('/', os.sep))
        if not os.path.exists(full_path):
            self.stdout.write(self.style.WARNING(f'    Image not found: {full_path}'))
            return None
        return File(open(full_path, 'rb'), name=os.path.basename(full_path))

    def _parse_date(self, val):
        if val is None or val == '':
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
        return None

    def _parse_decimal(self, val):
        if val is None or val == '':
            return None
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None

    def handle(self, *args, **options):
        self._ensure_migrated()
        self.stdout.write('Loading sample data from Excel files...')

        # ── Users ────────────────────────────────────────────────────
        self.stdout.write('Seeding users...')
        from apps.accounts.models import UserProfile
        users = {}
        for row in self._load_sheet('users.xlsx'):
            username = row.get('username')
            if not username:
                continue
            defaults = {
                'email': row.get('email', ''),
                'first_name': row.get('first_name', ''),
                'last_name': row.get('last_name', ''),
                'role': row.get('role', 'MEMBER'),
                'is_staff': row.get('role', 'MEMBER') == 'ADMIN',
                'is_superuser': username == 'tur0001',
            }
            user, created = User.objects.get_or_create(username=username, defaults=defaults)
            if created:
                user.set_password(row.get('password', 'changeme'))
                user.save()
                self.stdout.write(f'  Created user: {username}')
            else:
                self.stdout.write(f'  Using existing user: {username}')
            # Set avatar if image provided
            image_path = row.get('image_path', '')
            if image_path:
                profile = UserProfile.objects.get_or_create(user=user)[0]
                img_file = self._load_image(image_path)
                if img_file:
                    profile.avatar.save(os.path.basename(image_path), img_file, save=True)
                    self.stdout.write(f'    Set avatar for {username}')
            users[username] = user

        # ── Categories ───────────────────────────────────────────────
        self.stdout.write('Seeding categories...')
        from apps.equipment.models import Category
        categories = {}
        for row in self._load_sheet('categories.xlsx'):
            name = row.get('name')
            if not name:
                continue
            cat, created = Category.objects.get_or_create(name=name, defaults={
                'description': row.get('description', ''),
                'color': row.get('color', ''),
            })
            if created:
                cat.save()  # triggers auto-color generation if blank
                self.stdout.write(f'  Created category: {name}')
            categories[name] = cat

        # ── Locations ────────────────────────────────────────────────
        self.stdout.write('Seeding locations...')
        from apps.equipment.models import Location
        locations = {}
        for row in self._load_sheet('locations.xlsx'):
            name = row.get('name')
            if not name:
                continue
            loc, created = Location.objects.get_or_create(name=name, defaults={
                'description': row.get('description', ''),
                'building': row.get('building', ''),
                'room': row.get('room', ''),
            })
            if created:
                self.stdout.write(f'  Created location: {name}')
            locations[name] = loc

        # ── Equipment ────────────────────────────────────────────────
        self.stdout.write('Seeding equipment...')
        from apps.equipment.models import Equipment
        equipment_map = {}  # key by serial_number
        for row in self._load_sheet('equipment.xlsx'):
            serial = row.get('serial_number')
            if not serial:
                continue
            cat = categories.get(row.get('category_name'))
            loc = locations.get(row.get('location_name'))
            owner = users.get(row.get('owner_username'))
            defaults = {
                'name': row.get('name') or '',
                'description': row.get('description') or '',
                'model_number': row.get('model_number') or '',
                'manufacturer': row.get('manufacturer') or '',
                'category': cat,
                'location': loc,
                'owner': owner,
                'created_by': owner,
                'condition': row.get('condition') or 'GOOD',
                'status': row.get('status') or 'AVAILABLE',
                'approval_status': 'APPROVED',
                'purchase_date': self._parse_date(row.get('purchase_date')),
                'purchase_price': self._parse_decimal(row.get('purchase_price')),
                'notes': row.get('notes') or '',
            }
            eq, created = Equipment.objects.get_or_create(serial_number=serial, defaults=defaults)
            if created:
                # Auto-approve if creator == owner
                eq.approval_status = 'APPROVED'
                eq.save(update_fields=['approval_status'])
                self.stdout.write(f'  Created equipment: {eq.name}')
                img_path = row.get('image_path', '')
                if img_path:
                    img_file = self._load_image(img_path)
                    if img_file:
                        eq.image.save(os.path.basename(img_path), img_file, save=True)
                        self.stdout.write(f'    Set image for {eq.name}')
            else:
                self.stdout.write(f'  Using existing equipment: {eq.name}')
            equipment_map[serial] = eq

        # ── Kits ─────────────────────────────────────────────────────
        self.stdout.write('Seeding kits...')
        from apps.kits.models import Kit, KitItem
        kits = {}
        for row in self._load_sheet('kits.xlsx'):
            name = row.get('name')
            if not name:
                continue
            creator = users.get(row.get('created_by_username'))
            kit, created = Kit.objects.get_or_create(name=name, defaults={
                'description': row.get('description', ''),
                'created_by': creator,
            })
            if created:
                self.stdout.write(f'  Created kit: {name}')
            kits[name] = kit

        # ── Kit Items ────────────────────────────────────────────────
        self.stdout.write('Seeding kit items...')
        for row in self._load_sheet('kit_items.xlsx'):
            kit_name = row.get('kit_name')
            eq_serial = row.get('equipment_serial_number')
            qty = row.get('quantity', 1)
            if not kit_name or not eq_serial:
                continue
            kit = kits.get(kit_name)
            eq = equipment_map.get(eq_serial)
            if kit and eq:
                KitItem.objects.get_or_create(kit=kit, equipment=eq, defaults={'quantity': qty})
                self.stdout.write(f'  Added {eq.name} to {kit.name} (qty={qty})')

        # ── Reservations ─────────────────────────────────────────────
        self.stdout.write('Seeding reservations...')
        from apps.reservations.models import Reservation
        for row in self._load_sheet('reservations.xlsx'):
            requester = users.get(row.get('requester_username'))
            eq = equipment_map.get(row.get('equipment_serial_number', '')) or None
            kit = kits.get(row.get('kit_name', '')) or None
            start_date = self._parse_date(row.get('start_date'))
            end_date = self._parse_date(row.get('end_date'))
            if not requester or (not eq and not kit) or not start_date or not end_date:
                continue
            status = row.get('status', 'PENDING')
            res, created = Reservation.objects.get_or_create(
                requester=requester,
                equipment=eq,
                kit=kit,
                start_date=start_date,
                end_date=end_date,
                defaults={
                    'purpose': row.get('purpose', ''),
                    'status': status,
                }
            )
            if created:
                # Update equipment status for confirmed reservations
                if status == 'CONFIRMED':
                    if eq:
                        eq.status = 'AVAILABLE'
                        eq.save(update_fields=['status'])
                    elif kit:
                        for ki in kit.items.select_related('equipment'):
                            ki.equipment.status = 'AVAILABLE'
                            ki.equipment.save(update_fields=['status'])
                self.stdout.write(f'  Created reservation: {res}')


        # ── Incidents ────────────────────────────────────────────────
        self.stdout.write('Seeding incidents...')
        from apps.incidents.models import IncidentReport
        for row in self._load_sheet('incidents.xlsx'):
            eq = equipment_map.get(row.get('equipment_serial_number', ''))
            if not eq:
                continue
            reporter = users.get(row.get('reported_by_username', ''))
            assigned = users.get(row.get('assigned_to_username', '')) if row.get('assigned_to_username') else None
            incident, created = IncidentReport.objects.get_or_create(
                equipment=eq,
                title=row.get('title', ''),
                defaults={
                    'description': row.get('description', ''),
                    'severity': row.get('severity', 'MEDIUM'),
                    'status': row.get('status', 'OPEN'),
                    'reported_by': reporter,
                    'assigned_to': assigned,
                    'resolution': row.get('resolution', ''),
                }
            )
            if created:
                img_path = row.get('image_path', '')
                if img_path:
                    img_file = self._load_image(img_path)
                    if img_file:
                        incident.image.save(os.path.basename(img_path), img_file, save=True)
                self.stdout.write(f'  Created incident: {incident.title}')

        # ── Maintenance Logs ─────────────────────────────────────────
        self.stdout.write('Seeding maintenance logs...')
        from apps.incidents.models import MaintenanceLog
        for row in self._load_sheet('maintenance.xlsx'):
            eq = equipment_map.get(row.get('equipment_serial_number', ''))
            if not eq:
                continue
            performer = users.get(row.get('performed_by_username', '')) if row.get('performed_by_username') else None
            scheduled = self._parse_date(row.get('scheduled_date'))
            if not scheduled:
                continue
            ml, created = MaintenanceLog.objects.get_or_create(
                equipment=eq,
                maintenance_type=row.get('maintenance_type', 'PREVENTIVE'),
                scheduled_date=scheduled,
                defaults={
                    'status': row.get('status', 'SCHEDULED'),
                    'performed_by': performer,
                    'description': row.get('description', ''),
                    'completed_date': self._parse_date(row.get('completed_date')),
                    'cost': self._parse_decimal(row.get('cost')),
                    'notes': row.get('notes', ''),
                }
            )
            if created:
                self.stdout.write(f'  Created maintenance: {ml}')

        # ── Calibration Logs ─────────────────────────────────────────
        self.stdout.write('Seeding calibration logs...')
        from apps.incidents.models import CalibrationLog
        for row in self._load_sheet('calibration.xlsx'):
            eq = equipment_map.get(row.get('equipment_serial_number', ''))
            if not eq:
                continue
            calibrator = users.get(row.get('calibrated_by_username', '')) if row.get('calibrated_by_username') else None
            cal_date = self._parse_date(row.get('calibration_date'))
            if not cal_date:
                continue
            cl, created = CalibrationLog.objects.get_or_create(
                equipment=eq,
                calibration_date=cal_date,
                defaults={
                    'calibrated_by': calibrator,
                    'next_calibration_date': self._parse_date(row.get('next_calibration_date')),
                    'status': row.get('status', 'PENDING'),
                    'certificate_number': row.get('certificate_number', ''),
                    'notes': row.get('notes', ''),
                }
            )
            if created:
                self.stdout.write(f'  Created calibration: {cl}')

        self.stdout.write(self.style.SUCCESS('\nAll sample data loaded successfully!'))
        self.stdout.write('\nTest accounts:')
        self.stdout.write('  Admin:    tur0001@auburn.edu / admin123')
        self.stdout.write('  Members:  hhs0015@auburn.edu, mzw0147@auburn.edu, ... / member123')
